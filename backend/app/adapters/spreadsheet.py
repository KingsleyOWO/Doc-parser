"""Native spreadsheet parser for Excel-style inputs.

The parser preserves workbook/sheet/table structure before the document enters
MinerU-oriented normalization. It emits MinerU-compatible content_list items so
later pipeline stages can stay parser-agnostic.
"""

import html
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency is declared, guard keeps error readable
    xlrd = None


@dataclass
class SpreadsheetParseResult:
    success: bool
    output_dir: Path
    content_list_path: Path | None = None
    markdown_path: Path | None = None
    error: str | None = None
    stats: dict[str, Any] = field(default_factory=dict)


def parse_spreadsheet(input_path: Path, output_dir: Path) -> SpreadsheetParseResult:
    """Parse an XLS/XLSX workbook into MinerU-compatible content_list JSON."""
    output_root = output_dir / input_path.stem / "native"
    output_root.mkdir(parents=True, exist_ok=True)

    suffix = input_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx_workbook(input_path, output_root)
    if suffix == ".xls":
        return _parse_xls_workbook(input_path, output_root)

    return SpreadsheetParseResult(
        success=False,
        output_dir=output_root,
        error=f"Unsupported spreadsheet extension: {suffix}",
    )


def _parse_xlsx_workbook(input_path: Path, output_root: Path) -> SpreadsheetParseResult:
    try:
        workbook = load_workbook(input_path, data_only=True, read_only=False)
    except Exception as exc:
        return SpreadsheetParseResult(
            success=False,
            output_dir=output_root,
            error=f"Spreadsheet parse failed: {exc}",
        )

    try:
        sheets = [(sheet.title, _sheet_rows(sheet)) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    return _write_spreadsheet_outputs(input_path, output_root, sheets)


def _parse_xls_workbook(input_path: Path, output_root: Path) -> SpreadsheetParseResult:
    if xlrd is None:
        return SpreadsheetParseResult(
            success=False,
            output_dir=output_root,
            error="XLS parser dependency missing. Install xlrd>=2.0.1.",
        )

    try:
        workbook = xlrd.open_workbook(str(input_path), formatting_info=True)
    except Exception as exc:
        return SpreadsheetParseResult(
            success=False,
            output_dir=output_root,
            error=f"Spreadsheet parse failed: {exc}",
        )

    sheets = []
    for sheet in workbook.sheets():
        rows = _xls_sheet_rows(sheet)
        sheets.append((sheet.name, rows))

    return _write_spreadsheet_outputs(input_path, output_root, sheets)


def _write_spreadsheet_outputs(
    input_path: Path,
    output_root: Path,
    sheets: list[tuple[str, list[list[str]]]],
) -> SpreadsheetParseResult:
    content_items: list[dict[str, Any]] = []
    markdown_lines: list[str] = []
    table_count = 0
    non_empty_rows = 0

    for sheet_idx, (sheet_name, rows) in enumerate(sheets):
        if not rows:
            continue

        title = f"工作表：{sheet_name}"
        content_items.append(
            {
                "type": "text",
                "text": title,
                "text_level": 1,
                "page_idx": sheet_idx,
                "bbox": [0, 0, 0, 0],
            }
        )
        markdown_lines.append(f"# {title}")
        markdown_lines.append("")

        table_html = _rows_to_html_table(rows)
        caption = f"{sheet_name} 表格"
        content_items.append(
            {
                "type": "table",
                "table_caption": caption,
                "table_body": table_html,
                "page_idx": sheet_idx,
                "bbox": [0, 0, 0, 0],
                "metadata": {
                    "source": "native_spreadsheet",
                    "sheet_name": sheet_name,
                    "row_count": len(rows),
                    "column_count": max((len(row) for row in rows), default=0),
                },
            }
        )
        markdown_lines.append(table_html)
        markdown_lines.append("")
        table_count += 1
        non_empty_rows += len(rows)

    if not content_items:
        return SpreadsheetParseResult(
            success=False,
            output_dir=output_root,
            error="Spreadsheet contains no non-empty sheets",
        )

    content_list_path = output_root / f"{input_path.stem}_content_list.json"
    content_list_path.write_text(
        json.dumps(content_items, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    markdown_path = output_root / f"{input_path.stem}.md"
    markdown_path.write_text("\n".join(markdown_lines).strip() + "\n", encoding="utf-8")

    return SpreadsheetParseResult(
        success=True,
        output_dir=output_root,
        content_list_path=content_list_path,
        markdown_path=markdown_path,
        stats={
            "parser": "native_spreadsheet",
            "sheet_count": len(sheets),
            "table_count": table_count,
            "non_empty_rows": non_empty_rows,
        },
    )


def _xls_sheet_rows(sheet: Any) -> list[list[str]]:
    """Return non-empty XLS rows with merged-cell values expanded."""
    merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for row_start, row_end, col_start, col_end in getattr(sheet, "merged_cells", []):
        anchor = (row_start, col_start)
        for row_idx in range(row_start, row_end):
            for col_idx in range(col_start, col_end):
                merged_lookup[(row_idx, col_idx)] = anchor

    rows: list[list[str]] = []
    max_width = 0
    for row_idx in range(sheet.nrows):
        values: list[str] = []
        for col_idx in range(sheet.ncols):
            value = sheet.cell_value(row_idx, col_idx)
            anchor = merged_lookup.get((row_idx, col_idx))
            if anchor and anchor != (row_idx, col_idx):
                value = ""
            values.append(_format_cell_value(value))

        while values and not values[-1]:
            values.pop()
        if any(values):
            max_width = max(max_width, len(values))
            rows.append(values)

    return [row + [""] * (max_width - len(row)) for row in rows]


def _sheet_rows(sheet: Any) -> list[list[str]]:
    """Return non-empty used rows with merged-cell values expanded."""
    rows: list[list[str]] = []
    max_width = 0

    for row in sheet.iter_rows():
        values: list[str] = []
        for cell in row:
            value = None
            if isinstance(cell, MergedCell):
                value = None
            else:
                value = cell.value
            values.append(_format_cell_value(value))

        while values and not values[-1]:
            values.pop()
        if any(values):
            max_width = max(max_width, len(values))
            rows.append(values)

    normalized_rows = [row + [""] * (max_width - len(row)) for row in rows]
    return normalized_rows


def _merged_cell_lookup(sheet: Any) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for merged_range in sheet.merged_cells.ranges:
        anchor = merged_range.start_cell.coordinate
        for row in sheet.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                lookup[cell.coordinate] = anchor
    return lookup


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_to_html_table(rows: list[list[str]]) -> str:
    if not rows:
        return "<table></table>"

    lines = ["<table>"]
    header_idx = _header_row_index(rows)
    for row_idx, row in enumerate(rows):
        tag = "th" if row_idx == header_idx else "td"
        lines.append("  <tr>")
        for cell in row:
            lines.append(f"    <{tag}>{html.escape(cell)}</{tag}>")
        lines.append("  </tr>")
    lines.append("</table>")
    return "\n".join(lines)


def _header_row_index(rows: list[list[str]]) -> int:
    for idx, row in enumerate(rows[:10]):
        non_empty = [cell for cell in row if cell]
        if len(non_empty) >= 2:
            return idx
    return 0
